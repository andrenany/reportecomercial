# -*- coding: utf-8 -*-
"""
Rellena Costeo programas oficiales (ene–jul 2026) desde FVivaldi.
Solo recepción 2026 (fecha_recepcion). Solo lectura SQL.
"""
from __future__ import annotations

import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from generar_consulta_facturacion import cargar_credenciales, conectar

SRC = Path(r"c:\Users\jpaillaguala\Downloads\Costeo programas oficiales enero Julio 2026.xlsx")
OUT = Path(r"c:\Users\jpaillaguala\Downloads\Costeo programas oficiales enero Julio 2026_RELLENO.xlsx")
SHEET = 0  # primera hoja


def norm(s) -> str:
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    t = str(s).strip()
    t = unicodedata.normalize("NFKD", t)
    t = "".join(ch for ch in t if not unicodedata.combining(ch))
    t = re.sub(r"\s+", " ", t).upper()
    return t


# Excel empresa header (fila 2) -> match keys (substring)
EMPRESA_KEYS = {
    4: ["ECO SALMON"],
    8: ["INVERMAR"],
    12: ["AUSTRALIS MAR"],
    16: ["MULTI X"],
    20: ["MULTIEXPORT PACIFIC"],
    24: ["COOKE AQUACULTURE"],
    28: ["SALMONES BLUMAR S.A", "SALMONES BLUMAR SA"],  # no magallanes
    32: ["CAMANCHACA"],
    36: ["MOWI"],
    40: ["CHILCOS"],
    44: ["LAGO SOFIA", "LAGO SOFÍA"],
    48: ["BLUMAR MAGALLANES"],
    52: ["CALETA BAY MAR"],
    56: ["YADRAN", "YADR"],
    60: ["NALCAHUE"],
    64: ["AQUAGEN"],
    68: ["DALCAHUE"],
    74: ["ANTARTICA", "ANTÁRTICA"],
}

# (row excel) -> lista de nombres técnica FVivaldi (match exacto normalizado)
# También patrones contiene
TECNICA_MAP = {
    # Villarrica 4-21, P.Montt 24-41, Aysén 44-61 — misma técnica relativa
    "RT-PCR ISAV": ["RT-PCR ISAV"],
    "IDENTIFICACION ISAV SECUENCIAMIENTO": [
        "IDENTIFICACION HPR ISAV POR SECUENCIAMIENTO",
        "IDENTIFICACION ISAV SECUENCIAMIENTO",
    ],
    "RT-PCR ISAV HPR0": ["RT-PCR ISAV HPR0"],
    "PVA": [],  # sin match claro en FVivaldi
    "ELF": ["RT-PCR ELF"],
    "IFAT PISCIRICKETTSIA SALMONIS": ["IFAT PISCIRICKETTSIA SALMONIS"],
    "RT-PCR PISCIRICKETTSIA SALMONIS": ["RT-PCR PISCIRICKETTSIA SALMONIS"],
    "RT-PCR PROA": [],
    "VISITAS ELIMINACION": ["VISITA ASESORIA SANITARIA"],
    "CERTIFICACION PROA SALMON": ["CERTIFICACION PROA SALMON", "CERTIFICACION PROA SALMON PVE"],
    "CIM PISCIRICKETTSIA SALMONIS": ["CIM PISCIRICKETTSIA SALMONIS"],
    "CULTIVO ESPECIAL P. SALMONIS": [
        "CULTIVO ESPECIAL PISCIRICKETTSIA SALMONIS",
        "CULTIVO ESPECIAL P. SALMONIS",
    ],
    "IDENTIFICACION BACTERIANA PCR": [
        "IDENTIFICACION BACTERIANA MEDIANTE PCR",
        "IDENTIFICACION BACTERIANA PCR",
    ],
    "SERVICIO DE MUESTREO (ATL)": ["SERVICIO DE MUESTREO (ATL)"],
    "TOMA DE MUESTRA CLIENTE": ["TOMA DE MUESTRA CLIENTE"],
    "TOMA DE MUESTRA EN LABORATORIO": ["TOMA DE MUESTRA EN LABORATORIO"],
    "RT-PCR BKD AUSTRALIS MAGALLANES": [
        "RT-PCR RENIBACTERIUM SALMONINARUM",
        "RT-PCR BKD AUSTRALIS MAGALLANES",
        "RT-PCR BKD",
    ],
    "COSTO OPERATIVO": ["COSTO OPERATIVO"],
}

SEDE_ROWS = {
    "VILLARRICA": range(4, 22),
    "PUERTO MONTT": range(24, 42),
    "AYSEN": range(44, 62),
}

SEDE_ALIASES = {
    "VILLARRICA": ["VILLARRICA"],
    "PUERTO MONTT": ["PUERTO MONTT", "P. MONTT", "PTO MONTT"],
    "AYSEN": ["AYSEN", "AYSÉN", "AYSEN"],
}


def load_fv() -> pd.DataFrame:
    cfg = cargar_credenciales()
    conn = conectar(cfg)
    q = """
    SELECT
        ISNULL(NULLIF(LTRIM(RTRIM(nombre_lugaranalisis)), ''), '(sin sede)') AS sede,
        ISNULL(NULLIF(LTRIM(RTRIM(nombre_empresa)), ''), '(sin empresa)') AS empresa,
        ISNULL(NULLIF(LTRIM(RTRIM(nombre_tecnica)), ''), '(sin tecnica)') AS tecnica,
        ISNULL(NULLIF(LTRIM(RTRIM(nombre_seccion)), ''), '(sin seccion)') AS seccion,
        SUM(CAST(ISNULL(fac_n_analisis,0) AS FLOAT)) AS n_analisis,
        SUM(CAST(ISNULL(fac_vtatotal,0) AS FLOAT)) AS venta_uf
    FROM dbo.vw_FVivaldiWebSalud
    WHERE TRY_CONVERT(date, fecha_recepcion, 103) >= '2026-01-01'
      AND TRY_CONVERT(date, fecha_recepcion, 103) < '2027-01-01'
      AND TRY_CONVERT(date, fecha_recepcion, 103) IS NOT NULL
    GROUP BY
        ISNULL(NULLIF(LTRIM(RTRIM(nombre_lugaranalisis)), ''), '(sin sede)'),
        ISNULL(NULLIF(LTRIM(RTRIM(nombre_empresa)), ''), '(sin empresa)'),
        ISNULL(NULLIF(LTRIM(RTRIM(nombre_tecnica)), ''), '(sin tecnica)'),
        ISNULL(NULLIF(LTRIM(RTRIM(nombre_seccion)), ''), '(sin seccion)')
    """
    df = pd.read_sql(q, conn)
    conn.close()
    df["sede_n"] = df["sede"].map(norm)
    df["empresa_n"] = df["empresa"].map(norm)
    df["tecnica_n"] = df["tecnica"].map(norm)
    df["seccion_n"] = df["seccion"].map(norm)
    return df


def load_costos(path: Path) -> dict[str, float]:
    df = pd.read_excel(path, sheet_name="costos analisis", header=None)
    costos = {}
    for _, row in df.iterrows():
        tec = row[1]
        val = row[2]
        if pd.isna(tec) or pd.isna(val):
            continue
        try:
            costos[norm(tec)] = float(val)
        except (TypeError, ValueError):
            continue
    # aliases útiles
    alias = {
        "RT-PCR ISAV": costos.get("RT-PCR ISAV", 0.266),
        "IDENTIFICACION ISAV SECUENCIAMIENTO": costos.get(norm("Identificación ISAV secuenciamiento"), 0.89)
            or costos.get(norm("Identificacion ISAV secuenciamiento"), 0.89),
        "RT-PCR ISAV HPR0": costos.get("RT-PCR ISAV HPR0", 0.45),
        "PVA": costos.get(norm("Pva (pool 6 patogenos)"), 0.417),
        "ELF": costos.get("RT-PCR ELF", 0.09),
        "IFAT PISCIRICKETTSIA SALMONIS": costos.get(norm("IFAT Piscirickettsia salmonis"), 0.21),
        "RT-PCR PISCIRICKETTSIA SALMONIS": costos.get(norm("RT-PCR Piscirickettsia salmonis"), 0.275),
        "CIM PISCIRICKETTSIA SALMONIS": costos.get(norm("CIM Piscirickettsia salmonis"), 0.55),
        "CULTIVO ESPECIAL P. SALMONIS": costos.get(norm("Cultivo especial P. salmonis"), 0.44),
        "IDENTIFICACION BACTERIANA PCR": costos.get(norm("Identificación bacteriana mediante PCR"), 0.0),
    }
    costos.update(alias)
    return costos


def empresas_for_col(df: pd.DataFrame, col: int) -> list[str]:
    keys = EMPRESA_KEYS.get(col, [])
    out = []
    for emp in df["empresa_n"].unique():
        if col == 28:
            # Blumar continental: Blumar pero no Magallanes
            if "BLUMAR" in emp and "MAGALLANES" not in emp:
                out.append(emp)
            continue
        if any(k in emp for k in keys):
            out.append(emp)
    return out


def tecnicas_match(nombre_excel: str, df_tecnicas: set[str]) -> list[str]:
    key = norm(nombre_excel)
    aliases = TECNICA_MAP.get(key)
    if aliases is None:
        # buscar por clave flexible
        for k, v in TECNICA_MAP.items():
            if k in key or key in k:
                aliases = v
                break
    if not aliases:
        # exacto en FVivaldi
        if key in df_tecnicas:
            return [key]
        return []
    matched = []
    for a in aliases:
        an = norm(a)
        if an in df_tecnicas:
            matched.append(an)
            continue
        # contains
        for t in df_tecnicas:
            if an and (an == t or an in t or t in an):
                matched.append(t)
    return sorted(set(matched))


def sede_match(sede_excel: str, sede_n: str) -> bool:
    aliases = SEDE_ALIASES.get(sede_excel, [sede_excel])
    return any(a in sede_n or sede_n in a for a in aliases)


def main() -> None:
    print("Cargando FVivaldi recepción 2026...")
    fv = load_fv()
    print(f"  grupos: {len(fv):,} · UF: {fv['venta_uf'].sum():,.1f} · N: {fv['n_analisis'].sum():,.0f}")
    costos = load_costos(SRC)
    tec_set = set(fv["tecnica_n"].unique())

    wb = load_workbook(SRC)
    ws = wb.worksheets[SHEET]

    # Detect technique name per row from col C (3)
    report = []
    filled = 0

    for sede_name, rows in SEDE_ROWS.items():
        for r in rows:
            tec_excel = ws.cell(r + 1, 3).value  # openpyxl 1-based; dataframe was 0-based row r col 2
            # Wait: pandas iloc[r,2] is row r; openpyxl row = r+1
            if tec_excel is None:
                continue
            tec_label = str(tec_excel).strip()
            if tec_label.lower() == "total":
                continue

            is_costo_op = "COSTO OPERATIVO" in norm(tec_label)
            matched_tecs = tecnicas_match(tec_label, tec_set)

            # unit cost
            unit = costos.get(norm(tec_label), 0.0)
            if not unit:
                for k, v in costos.items():
                    if norm(tec_label) in k or k in norm(tec_label):
                        unit = v
                        break

            for emp_col, keys in EMPRESA_KEYS.items():
                emps = empresas_for_col(fv, emp_col)
                if not emps and keys:
                    # try again loose
                    for emp in fv["empresa_n"].unique():
                        if any(k in emp for k in keys):
                            if emp_col == 28 and "MAGALLANES" in emp:
                                continue
                            emps.append(emp)

                mask = fv["empresa_n"].isin(emps) & fv["sede_n"].map(lambda s: sede_match(sede_name, s))
                if is_costo_op:
                    sub = fv[mask & (fv["seccion_n"] == "COSTO OPERATIVO")]
                    n = float(sub["n_analisis"].sum())
                    uf = float(sub["venta_uf"].sum())
                    cost = uf  # el costeo operativo ya viene en UF facturado
                else:
                    if not matched_tecs:
                        n, uf, cost = 0.0, 0.0, 0.0
                    else:
                        sub = fv[mask & fv["tecnica_n"].isin(matched_tecs)]
                        # evitar doble contar costo operativo en técnicas normales
                        sub = sub[sub["seccion_n"] != "COSTO OPERATIVO"]
                        n = float(sub["n_analisis"].sum())
                        uf = float(sub["venta_uf"].sum())
                        cost = n * float(unit or 0)

                # columnas: analisis=emp_col, facturado=+1, costos=+2 (0-based) → openpyxl +1
                c_n = emp_col + 1
                c_uf = emp_col + 2
                c_cost = emp_col + 3
                row_xl = r + 1

                if n or uf or (is_costo_op and (n or uf)):
                    ws.cell(row_xl, c_n).value = int(round(n))
                    ws.cell(row_xl, c_uf).value = round(uf, 3)
                    ws.cell(row_xl, c_cost).value = round(cost, 3)
                    filled += 1
                    report.append({
                        "sede": sede_name,
                        "tecnica_excel": tec_label,
                        "match_fv": "; ".join(matched_tecs) if matched_tecs else ("COSTO OPERATIVO seccion" if is_costo_op else "(sin match)"),
                        "empresa_col": emp_col,
                        "n": n,
                        "uf": uf,
                        "costo": cost,
                    })

    # Totals por bloque sede: fila Total
    for sede_name, rows in SEDE_ROWS.items():
        total_row = None
        for r in rows:
            if str(ws.cell(r + 1, 3).value or "").strip().lower() == "total":
                total_row = r + 1
                break
        if not total_row:
            continue
        data_rows = [r + 1 for r in rows if r + 1 != total_row]
        for emp_col in EMPRESA_KEYS:
            for offset in (0, 1, 2):
                c = emp_col + 1 + offset
                s = 0.0
                any_val = False
                for rr in data_rows:
                    v = ws.cell(rr, c).value
                    if isinstance(v, (int, float)) and not isinstance(v, bool):
                        s += float(v)
                        any_val = True
                if any_val:
                    ws.cell(total_row, c).value = int(round(s)) if offset == 0 else round(s, 3)

    # Columna cantidad análisis / costos / dif (78,79,80) — totales por técnica fila
    # pandas cols 78,79 = Cantidad analisis, costos
    for sede_name, rows in SEDE_ROWS.items():
        for r in rows:
            row_xl = r + 1
            tec = ws.cell(row_xl, 3).value
            if tec is None or str(tec).strip().lower() == "total":
                continue
            tot_n = 0.0
            tot_c = 0.0
            for emp_col in EMPRESA_KEYS:
                vn = ws.cell(row_xl, emp_col + 1).value
                vc = ws.cell(row_xl, emp_col + 3).value
                if isinstance(vn, (int, float)):
                    tot_n += float(vn)
                if isinstance(vc, (int, float)):
                    tot_c += float(vc)
            if tot_n or tot_c:
                ws.cell(row_xl, 79).value = int(round(tot_n))
                ws.cell(row_xl, 80).value = round(tot_c, 3)

    wb.save(OUT)
    rep = pd.DataFrame(report)
    rep_path = OUT.with_suffix(".csv")
    rep.to_csv(rep_path, index=False, encoding="utf-8-sig")
    print(f"OK -> {OUT}")
    print(f"Celdas con dato: {filled}")
    print(f"Detalle -> {rep_path}")
    sin = [t for t, aliases in TECNICA_MAP.items() if not aliases]
    if sin:
        print("Sin match FVivaldi (quedan vacías):", ", ".join(sin))


if __name__ == "__main__":
    main()
