"""
Lee dos fuentes de red (solo lectura; no modifica los Excel):

1) OC Pendientes de facturar.xlsx  -> hoja OC PENDIENTES
2) Gráficos Facturación 2026 ... Juan original.xlsx
     -> KR Resumen Consolidado (facturación)
     -> Tablas Auxiliares (RUT / Empresa / Empresa_Abr)

En memoria:
- color de fila (col A) -> estado de negocio
- limpia fechas
- normaliza clientes/programas
- aplica alias comercial Empresa_Abr cuando hay match
"""

from __future__ import annotations

import re
import unicodedata
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

# --- Fuentes (solo lectura) ---
EXCEL_OC = r"\\192.168.10.5\ws2016\Compartir Administracion\OC Pendientes de facturar.xlsx"
HOJA_OC = "OC PENDIENTES"

EXCEL_FAC = (
    r"\\192.168.10.5\adl.ws\Disco I\PM\COM\Carpeta compartida comercial"
    r"\analsisi fanny\Gráficos Facturación 2026 220126 Juan original.xlsx"
)
HOJA_FAC = "KR Resumen Consolidado"
HOJA_AUX = "Tablas Auxiliares"

# Facturado 2026.xlsx: ya no se usa en Solo facturación (fuente = Juan).
# Se deja el path por si hace falta un cruce puntual a futuro.
EXCEL_FACTURADO_2026 = (
    r"\\192.168.10.5\adl.ws\Disco I\PM\COM\Carpeta compartida comercial"
    r"\analsisi fanny\facturado 2026.xlsx"
)
HOJA_FACTURADO_2026 = "2026"

DIR_SALIDA = Path(__file__).resolve().parent

# Columnas útiles de OC PENDIENTES (A–T)
COLUMNAS_OC = [
    "ID",
    "Tipo venta",
    "Cliente",
    "Fecha",
    "AÑO",
    "Mes",
    "col_G",
    "uf_totalotrosdes",
    "uf_totalcaso",
    "uf_totalfac",
    "uf",
    "Total $",
    "Programa",
    "Observaciones",
    "N° OC",
    "N° HES",
    "N° FV",
    "Fecha_doc",
    "Usuario",
    "AÑO_doc",
]

THEME_RGB = {
    0: "FF000000",
    1: "FFFFFFFF",
    2: "FFEEECE1",
    3: "FF1F497D",
    4: "FF4F81BD",
    5: "FFC0504D",
    6: "FF9BBB59",
    7: "FF8064A2",
    8: "FF4BACC6",
    9: "FFF79646",
    10: "FF0000FF",
}

# Leyenda OC (columna A)
# sin color -> pendiente_facturar | azul -> listo_para_facturar
# verde -> facturada_ok | rojo/naranja -> anulada
COLOR_A_ESTADO = {
    None: ("sin_color", "pendiente_facturar"),
    "FFFFFFFF": ("sin_color", "pendiente_facturar"),
    "FF4F81BD": ("azul", "listo_para_facturar"),
    "FF0000FF": ("azul", "listo_para_facturar"),
    "FFA9D08E": ("verde", "facturada_ok"),
    "FF92D050": ("verde", "facturada_ok"),
    "FFFF2929": ("rojo", "anulada"),
    "FFFF0000": ("rojo", "anulada"),
    "FFF79646": ("naranja", "anulada"),
    "FFFFFF00": ("amarillo", "sin_leyenda"),
}


def _color_celda(cell) -> str | None:
    fill = cell.fill
    if fill is None or not fill.fill_type or fill.fill_type == "none":
        return None
    fg = fill.fgColor
    if fg is None:
        return None
    tipo = getattr(fg, "type", None)
    if tipo == "rgb":
        rgb = fg.rgb
        if isinstance(rgb, str) and len(rgb) >= 6:
            return rgb.upper()
    if tipo == "theme" and fg.theme is not None:
        return THEME_RGB.get(int(fg.theme), f"THEME_{fg.theme}")
    if tipo == "indexed" and fg.indexed is not None:
        return f"INDEXED_{fg.indexed}"
    return None


def _color_y_estado(rgb: str | None) -> tuple[str, str]:
    if rgb in COLOR_A_ESTADO:
        return COLOR_A_ESTADO[rgb]
    if rgb and rgb.startswith("FF4F81"):
        return ("azul", "listo_para_facturar")
    return (rgb or "desconocido", "sin_leyenda")


def _quitar_acentos(texto: str) -> str:
    nfkd = unicodedata.normalize("NFKD", texto)
    return "".join(ch for ch in nfkd if not unicodedata.combining(ch))


def normalizar_texto(valor) -> str | None:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    texto = str(valor).strip()
    if not texto or texto.lower() == "nan":
        return None
    texto = re.sub(r"\s+", " ", texto)
    texto = _quitar_acentos(texto).upper()
    texto = texto.replace(" S.A.", " SA").replace(" S.A", " SA")
    texto = texto.replace(" LTDA.", " LTDA").replace(" LIMITADA", " LTDA")
    texto = texto.replace(" SPA.", " SPA")
    return texto


def normalizar_rut(valor) -> str | None:
    if valor is None or (isinstance(valor, float) and pd.isna(valor)):
        return None
    texto = str(valor).strip().upper().replace(".", "").replace(" ", "")
    if not texto or texto == "NAN":
        return None
    return texto


def limpiar_fecha(serie: pd.Series) -> pd.Series:
    dt = pd.to_datetime(serie, errors="coerce", dayfirst=True)
    mask = dt.isna() & serie.notna()
    if mask.any():
        texto = serie.loc[mask].astype(str).str.strip()
        texto = texto.str.replace(r"\s+\d{1,2}:\d{2}:\d{2}.*$", "", regex=True)
        dt.loc[mask] = pd.to_datetime(texto, errors="coerce", dayfirst=True)
    return dt


def normalizar_sede(valor) -> str | None:
    t = normalizar_texto(valor)
    if t is None:
        return None
    mapa = {
        "PUERTO MONTT": "PUERTO MONTT",
        "VILLARRICA": "VILLARRICA",
        "AYSEN": "AYSEN",
        "NO CORRESPONDE": "NO CORRESPONDE",
    }
    return mapa.get(t, t)


def sede_desde_cuenta(cuenta) -> str | None:
    """Infiere sede solo cuando la cuenta contable es claramente geográfica."""
    cu = normalizar_texto(cuenta) or ""
    if not cu:
        return None
    if "VILLARRICA" in cu:
        return "VILLARRICA"
    if "AYSEN" in cu or "AYSEN" in cu.replace("É", "E"):
        return "AYSEN"
    if "MONTT" in cu or "PTO MONTT" in cu or "PUERTO MONTT" in cu:
        return "PUERTO MONTT"
    # Cuentas de producto/programa (ISAV, PVA, Screening, Asist., Facturas por Emitir, etc.)
    # no definen sede: debe venir del consolidado Juan.
    return None


# ---------------------------------------------------------------------------
# Lectura OC Pendientes
# ---------------------------------------------------------------------------

def leer_oc_pendientes(path: str = EXCEL_OC) -> pd.DataFrame:
    wb = load_workbook(path, data_only=True, read_only=False)
    ws = wb[HOJA_OC]
    filas = []
    max_col = min(ws.max_column, len(COLUMNAS_OC))

    for r in range(2, ws.max_row + 1):
        valores = [ws.cell(r, c).value for c in range(1, max_col + 1)]
        if all(v is None or (isinstance(v, str) and not v.strip()) for v in valores):
            continue
        rgb = _color_celda(ws.cell(r, 1))
        color, estado = _color_y_estado(rgb)
        registro = dict(zip(COLUMNAS_OC[:max_col], valores))
        registro["fila_excel"] = r
        registro["color_rgb"] = rgb
        registro["color"] = color
        registro["estado"] = estado
        filas.append(registro)

    wb.close()
    return pd.DataFrame(filas)


# ---------------------------------------------------------------------------
# Lectura Facturación Juan + tablas auxiliares
# ---------------------------------------------------------------------------

def leer_mapa_empresas(path: str = EXCEL_FAC) -> pd.DataFrame:
    """RUT / razón social / nombre corto comercial (Empresa_Abr)."""
    aux = pd.read_excel(path, sheet_name=HOJA_AUX, header=None)
    # Bloque izquierdo: Rut | Empresa | Empresa_Abr (filas con header en 0)
    block = aux.iloc[:, 0:3].copy()
    block.columns = ["Rut", "Empresa", "Empresa_Abr"]
    block = block.iloc[1:].dropna(how="all")
    block["Rut_norm"] = block["Rut"].map(normalizar_rut)
    block["Empresa_norm"] = block["Empresa"].map(normalizar_texto)
    block["Empresa_Abr_norm"] = block["Empresa_Abr"].map(normalizar_texto)
    block = block.dropna(subset=["Empresa_norm"]).drop_duplicates(subset=["Empresa_norm"])
    return block.reset_index(drop=True)


def leer_facturacion(path: str = EXCEL_FAC) -> pd.DataFrame:
    """Hoja KR Resumen Consolidado (detalle de documentos facturados)."""
    df = pd.read_excel(path, sheet_name=HOJA_FAC)
    df.columns = [str(c).strip() for c in df.columns]

    # Renombres estables
    rename = {
        "NOMBRE": "Cliente",
        " GLOSA": "Glosa",
        "GLOSA": "Glosa",
        "NUM. DOC.": "Num_Doc",
        "N° CUENTA": "Num_Cuenta",
        "CUENTA": "Cuenta",
        "TIPO DOC": "Tipo_Doc",
        "Tipo Ingreso": "Tipo_Ingreso",
        "Check Acumulado": "Check_Acumulado",
        "Año": "Anio",
        "Ano": "Anio",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    # Año puede venir con encoding raro (Año / A�o)
    if "Anio" not in df.columns:
        for c in list(df.columns):
            key = re.sub(r"[^a-z]", "", c.lower().replace("ñ", "n").replace("�", "n"))
            if key == "ano":
                df = df.rename(columns={c: "Anio"})
                break

    df["MONTO"] = pd.to_numeric(df["MONTO"], errors="coerce")
    df["Anio"] = pd.to_numeric(df["Anio"], errors="coerce").astype("Int64")
    if "Num_Doc" in df.columns:
        df["Num_Doc"] = pd.to_numeric(df["Num_Doc"], errors="coerce")
    df["Cliente_norm"] = df["Cliente"].map(normalizar_texto)
    df["Empresa_norm"] = df["Empresa"].map(normalizar_texto)
    df["Sede_norm"] = df["Sede"].map(normalizar_sede)
    df["Tipo_Ingreso_norm"] = df["Tipo_Ingreso"].map(normalizar_texto)
    df["Rut_norm"] = df["RUT"].map(normalizar_rut)
    df["PERIODO_norm"] = df["PERIODO"].map(normalizar_texto)
    return df


def leer_facturado_2026(path: str = EXCEL_FACTURADO_2026) -> pd.DataFrame:
    """
    Listado de líneas facturadas en 2026 (archivo facturado 2026.xlsx).
    Conserva todas las filas (un doc puede tener varias líneas/cuentas).
    """
    df = pd.read_excel(path, sheet_name=HOJA_FACTURADO_2026)
    df.columns = [str(c).strip() for c in df.columns]
    rename = {
        "NOMBRE": "Cliente",
        "NOMBRE ": "Cliente",
        "GLOSA": "Glosa",
        " GLOSA": "Glosa",
        "NUM. DOC.": "Num_Doc",
        "NUM. DOC. ": "Num_Doc",
        "CUENTA": "Cuenta",
        "CUENTA ": "Cuenta",
        "N° CUENTA": "Num_Cuenta",
        "TIPO DOC": "Tipo_Doc",
    }
    df = df.rename(columns={k: v for k, v in rename.items() if k in df.columns})
    if "Num_Doc" not in df.columns:
        for c in df.columns:
            if "DOC" in c.upper() and "TIPO" not in c.upper():
                df = df.rename(columns={c: "Num_Doc"})
                break
    if "Cliente" not in df.columns:
        for c in df.columns:
            if c.upper().startswith("NOMBRE"):
                df = df.rename(columns={c: "Cliente"})
                break
    if "Glosa" not in df.columns:
        for c in df.columns:
            if "GLOSA" in c.upper():
                df = df.rename(columns={c: "Glosa"})
                break
    if "Cuenta" not in df.columns:
        for c in df.columns:
            if c.upper().startswith("CUENTA"):
                df = df.rename(columns={c: "Cuenta"})
                break

    df["Num_Doc"] = pd.to_numeric(df["Num_Doc"], errors="coerce")
    df["MONTO"] = pd.to_numeric(df.get("MONTO"), errors="coerce")
    df["PERIODO_dt"] = pd.to_datetime(df.get("PERIODO"), errors="coerce")
    df["anio_factura"] = df["PERIODO_dt"].dt.year.astype("Int64")
    df["mes_factura"] = df["PERIODO_dt"].dt.month.astype("Int64")
    df["Cliente_norm"] = df["Cliente"].map(normalizar_texto) if "Cliente" in df.columns else None
    if "RUT" in df.columns:
        df["Rut_norm"] = df["RUT"].map(normalizar_rut)
    else:
        df["Rut_norm"] = None
    cuenta = df["Cuenta"] if "Cuenta" in df.columns else pd.Series([None] * len(df))
    df["Sede_norm"] = cuenta.map(sede_desde_cuenta)
    df = df.dropna(subset=["Num_Doc"])
    return df.reset_index(drop=True)


def aplicar_alias_empresa(df: pd.DataFrame, mapa: pd.DataFrame, col_cliente: str = "Cliente_norm") -> pd.DataFrame:
    """Agrega Empresa_Abr a partir del nombre de cliente (match por razón social)."""
    out = df.copy()
    lookup = (
        mapa.dropna(subset=["Empresa_norm", "Empresa_Abr_norm"])
        .drop_duplicates("Empresa_norm")
        .set_index("Empresa_norm")["Empresa_Abr_norm"]
    )
    out["Empresa_Abr"] = out[col_cliente].map(lookup)
    # Si no hay match por razón social, usa el propio nombre normalizado
    out["Cliente_key"] = out["Empresa_Abr"].fillna(out[col_cliente])
    return out


def preparar_oc(df: pd.DataFrame, mapa: pd.DataFrame | None = None) -> pd.DataFrame:
    out = df.copy()
    out["Fecha_limpia"] = limpiar_fecha(out["Fecha"])
    out["Cliente_norm"] = out["Cliente"].map(normalizar_texto)
    out["Programa_norm"] = out["Programa"].map(normalizar_texto)
    out["Total_num"] = pd.to_numeric(out["Total $"], errors="coerce")
    if mapa is not None:
        out = aplicar_alias_empresa(out, mapa, "Cliente_norm")
    else:
        out["Empresa_Abr"] = None
        out["Cliente_key"] = out["Cliente_norm"]
    return out


def preparar_fac(df: pd.DataFrame, mapa: pd.DataFrame | None = None) -> pd.DataFrame:
    out = df.copy()
    if mapa is not None:
        # Preferir match por RUT; si no, por razón social
        out = out.copy()
        by_rut = (
            mapa.dropna(subset=["Rut_norm", "Empresa_Abr_norm"])
            .drop_duplicates("Rut_norm")
            .set_index("Rut_norm")["Empresa_Abr_norm"]
        )
        by_nom = (
            mapa.dropna(subset=["Empresa_norm", "Empresa_Abr_norm"])
            .drop_duplicates("Empresa_norm")
            .set_index("Empresa_norm")["Empresa_Abr_norm"]
        )
        out["Empresa_Abr"] = out["Rut_norm"].map(by_rut)
        out["Empresa_Abr"] = out["Empresa_Abr"].fillna(out["Cliente_norm"].map(by_nom))
        out["Empresa_Abr"] = out["Empresa_Abr"].fillna(out["Empresa_norm"])
        out["Cliente_key"] = out["Empresa_Abr"]
    else:
        out["Empresa_Abr"] = out["Empresa_norm"]
        out["Cliente_key"] = out["Empresa_norm"]
    return out


def main() -> None:
    pd.set_option("display.max_columns", 20)
    pd.set_option("display.width", 220)
    pd.set_option("display.max_colwidth", 35)
    pd.set_option("display.float_format", lambda x: f"{x:,.0f}")

    print("Leyendo tablas auxiliares (alias empresas)...")
    mapa = leer_mapa_empresas()
    print(f"  Alias empresas: {len(mapa):,}")

    print(f"\nLeyendo OC Pendientes:\n  {EXCEL_OC}\n  hoja: {HOJA_OC}")
    oc = preparar_oc(leer_oc_pendientes(), mapa)
    print(f"  Filas: {len(oc):,}")

    print(f"\nLeyendo Facturación:\n  {EXCEL_FAC}\n  hoja: {HOJA_FAC}")
    fac = preparar_fac(leer_facturacion(), mapa)
    print(f"  Filas: {len(fac):,}")

    print("\n=== OC: estados ===")
    print(
        oc.groupby(["estado", "color"], dropna=False)
        .agg(filas=("ID", "count"), total_clp=("Total_num", "sum"))
        .sort_values("filas", ascending=False)
        .to_string()
    )

    print("\n=== Facturación: totales por año ===")
    print(
        fac.groupby("Anio")["MONTO"]
        .agg(docs="count", total_clp="sum")
        .to_string()
    )

    print("\n=== Facturación 2026: top 10 por Empresa_Abr ===")
    top = (
        fac.loc[fac["Anio"] == 2026]
        .groupby("Cliente_key")["MONTO"]
        .sum()
        .sort_values(ascending=False)
        .head(10)
    )
    print(top.to_string())

    print("\n=== Facturación 2026: por Tipo_Ingreso ===")
    print(
        fac.loc[fac["Anio"] == 2026]
        .groupby("Tipo_Ingreso_norm")["MONTO"]
        .agg(docs="count", total_clp="sum")
        .sort_values("total_clp", ascending=False)
        .to_string()
    )

    # Export local (no toca Excel de red)
    oc_out = DIR_SALIDA / "oc_pendientes_lectura.csv"
    fac_out = DIR_SALIDA / "facturacion_consolidado_lectura.csv"
    mapa_out = DIR_SALIDA / "mapa_empresas.csv"
    oc.to_csv(oc_out, index=False, encoding="utf-8-sig")
    fac.to_csv(fac_out, index=False, encoding="utf-8-sig")
    mapa.to_csv(mapa_out, index=False, encoding="utf-8-sig")

    print(f"\nCopias de trabajo:")
    print(f"  {oc_out}")
    print(f"  {fac_out}")
    print(f"  {mapa_out}")
    print("(Los Excel de red no fueron modificados.)")


if __name__ == "__main__":
    main()
